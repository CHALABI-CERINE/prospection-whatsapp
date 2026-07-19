#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Recherche d'établissements par ville et catégorie, avec ou sans site web.

Source de données : OpenStreetMap (via Nominatim pour géocoder la ville,
et Overpass API pour interroger les établissements). Base de données
ouverte et gratuite, avec une API publique officielle — donc pas de
scraping de Google Maps ni de violation de conditions d'utilisation.
Le lien "Google Maps" fourni dans les résultats n'est qu'un lien de
consultation (https://www.google.com/maps?q=lat,lon).

Fonctionnalités :
- Suggestions de ville en direct pendant la saisie (filtrées par le
  pays/état indiqué), pour choisir rapidement la bonne ville sans
  fenêtre de confirmation qui bloque le travail.
- Plusieurs serveurs Overpass essayés automatiquement en cas d'erreur
  (504 / indisponibilité) pour fiabiliser la recherche.
- Trois modes de recherche :
    * Sans site web : prospection — les établissements qui n'ont AUCUN
      moyen de contact (ni téléphone, ni e-mail) sont exclus, car
      injoignables.
    * Avec site web : établissements à qui proposer d'autres
      fonctionnalités sur leur site existant, ou des mesures de
      sécurisation.
    * Les deux à la fois.
- Export des résultats en fichier Excel (.xlsx).

Dépendances (à installer une seule fois) :
    pip install requests openpyxl

Lancement :
    python app_sans_site_web.py
"""

import threading
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import webbrowser

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration des APIs (gratuites, sans clé)
# ---------------------------------------------------------------------------
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"

# Plusieurs serveurs Overpass publics : si l'un est en erreur 504 /
# indisponible, on essaie automatiquement le suivant.
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.openstreetmap.fr/api/interpreter",
    "https://overpass.private.coffee/api/interpreter",
]

# Nominatim et Overpass demandent un User-Agent identifiable (politique
# d'usage des projets OpenStreetMap) — on en fournit un.
HEADERS = {"User-Agent": "RechercheSansSiteWeb/1.0 (usage personnel)"}

# Correspondance "catégorie en français" -> tags OpenStreetMap (clé, valeur)
CATEGORIES = {
    "Marché / Market": [("amenity", "marketplace"), ("shop", "marketplace")],
    "Supermarché": [("shop", "supermarket")],
    "Épicerie": [("shop", "convenience"), ("shop", "grocery")],
    "Pharmacie": [("amenity", "pharmacy")],
    "Restaurant": [("amenity", "restaurant")],
    "Café": [("amenity", "cafe")],
    "Boulangerie": [("shop", "bakery")],
    "Boucherie": [("shop", "butcher")],
    "Coiffeur": [("shop", "hairdresser")],
    "Garage / Mécanicien": [("shop", "car_repair")],
    "Hôtel": [("tourism", "hotel")],
    "Banque": [("amenity", "bank")],
    "Vêtements": [("shop", "clothes")],
    "Électronique": [("shop", "electronics")],
    "Quincaillerie": [("shop", "hardware")],
}

# Modes de recherche disponibles dans l'interface
MODE_SANS_SITE = "Sans site web (prospection)"
MODE_AVEC_SITE = "Avec site web (améliorations / sécurisation)"
MODE_LES_DEUX = "Les deux"
MODE_CODES = {
    MODE_SANS_SITE: "no_website",
    MODE_AVEC_SITE: "with_website",
    MODE_LES_DEUX: "both",
}


def guess_tags_from_free_text(text):
    """Si l'utilisateur tape une catégorie non listée, on tente une
    correspondance directe sur shop=<texte> (comportement de repli)."""
    text = text.strip().lower().replace(" ", "_")
    return [("shop", text)]


# ---------------------------------------------------------------------------
# Géocodage (recherche de ville + suggestions)
# ---------------------------------------------------------------------------
class SearchError(Exception):
    pass


def geocode_candidates(city_name, region_name=None, limit=6):
    """Retourne une liste de correspondances Nominatim pour la ville
    (et le pays/état si précisé). Utilisé aussi bien pour les
    suggestions en direct que pour le géocodage final."""
    city_name = city_name.strip()
    region_name = (region_name or "").strip()
    if not city_name:
        return []

    queries = []
    if region_name:
        queries.append(f"{city_name}, {region_name}")
    queries.append(city_name)

    for query in queries:
        params = {
            "q": query,
            "format": "json",
            "limit": limit,
            "addressdetails": 1,
        }
        resp = requests.get(NOMINATIM_URL, params=params, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data:
            return data

    return []


def geocode_best_match(city_name, region_name=None):
    """Renvoie directement la meilleure correspondance (première ville
    trouvée), sans fenêtre de confirmation — utilisé quand l'utilisateur
    n'a pas sélectionné de suggestion dans la liste déroulante."""
    candidates = geocode_candidates(city_name, region_name, limit=1)
    if not candidates:
        raise SearchError(f"Ville introuvable : {city_name}")
    return candidates[0]


# ---------------------------------------------------------------------------
# Overpass : requête et extraction des établissements
# ---------------------------------------------------------------------------
def build_overpass_query(geo, tag_pairs):
    """Construit une requête Overpass QL. Utilise la zone administrative
    (area) si la ville correspond à une relation OSM ; sinon, se replie
    sur une recherche dans le rectangle englobant (bounding box)."""
    tag_filters = "".join(
        f'node["{k}"="{v}"](area.searchArea);\n  way["{k}"="{v}"](area.searchArea);\n  relation["{k}"="{v}"](area.searchArea);\n  '
        for k, v in tag_pairs
    )

    if geo.get("osm_type") == "relation":
        area_id = 3600000000 + int(geo["osm_id"])
        return f"""
[out:json][timeout:60];
area({area_id})->.searchArea;
(
  {tag_filters}
);
out center tags;
"""

    south, north, west, east = geo["boundingbox"][0], geo["boundingbox"][1], geo["boundingbox"][2], geo["boundingbox"][3]
    bbox = f"{south},{west},{north},{east}"
    bbox_filters = "".join(
        f'node["{k}"="{v}"]({bbox});\n  way["{k}"="{v}"]({bbox});\n  '
        for k, v in tag_pairs
    )
    return f"""
[out:json][timeout:60];
(
  {bbox_filters}
);
out center tags;
"""


def run_overpass(query, progress_callback=None):
    """Essaie chaque serveur Overpass à tour de rôle. En cas d'erreur
    serveur (504, 502, 503...), de timeout ou de réseau, passe au
    serveur suivant plutôt que d'afficher une erreur immédiate."""
    last_error = None
    for i, url in enumerate(OVERPASS_ENDPOINTS):
        try:
            if progress_callback and i > 0:
                progress_callback(f"Serveur précédent indisponible, nouvel essai ({i + 1}/{len(OVERPASS_ENDPOINTS)})...")
            resp = requests.post(url, data={"data": query}, headers=HEADERS, timeout=75)
            if resp.status_code == 429:
                last_error = SearchError("Trop de requêtes envoyées, réessayez dans une minute.")
                continue
            if resp.status_code >= 500:
                last_error = SearchError(f"Serveur Overpass indisponible ({resp.status_code}).")
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as exc:
            last_error = SearchError(f"Erreur réseau sur un serveur Overpass : {exc}")
            continue

    raise last_error or SearchError(
        "Tous les serveurs Overpass sont indisponibles pour le moment. Réessayez dans quelques minutes."
    )


def has_website(tags):
    return any(tags.get(k) for k in ("website", "contact:website", "url"))


def has_phone(tags):
    return any(tags.get(k) for k in ("phone", "contact:phone", "contact:mobile"))


def has_email(tags):
    return any(tags.get(k) for k in ("email", "contact:email"))


def extract_address(tags):
    parts = []
    if tags.get("addr:housenumber") and tags.get("addr:street"):
        parts.append(f"{tags['addr:housenumber']} {tags['addr:street']}")
    elif tags.get("addr:street"):
        parts.append(tags["addr:street"])
    if tags.get("addr:city"):
        parts.append(tags["addr:city"])
    if tags.get("addr:postcode"):
        parts.append(tags["addr:postcode"])
    return ", ".join(parts) if parts else "Non disponible"


def extract_phone(tags):
    for key in ("phone", "contact:phone", "contact:mobile"):
        if tags.get(key):
            return tags[key]
    return "Non disponible"


def extract_email(tags):
    for key in ("email", "contact:email"):
        if tags.get(key):
            return tags[key]
    return "Non disponible"


def extract_website(tags):
    for key in ("website", "contact:website", "url"):
        if tags.get(key):
            return tags[key]
    return "Aucun"


def search_places(geo, category_label, mode="no_website", progress_callback=None):
    """Interroge Overpass pour la zone déjà géocodée (geo) et renvoie la
    liste des établissements correspondant au mode choisi :
      - "no_website" : uniquement ceux SANS site web, en excluant ceux
        qui n'ont aucun moyen de contact (ni téléphone, ni e-mail).
      - "with_website" : uniquement ceux qui ONT un site web.
      - "both" : les deux catégories, avec la colonne "Statut" pour les
        distinguer.
    """
    tag_pairs = CATEGORIES.get(category_label)
    if tag_pairs is None:
        tag_pairs = guess_tags_from_free_text(category_label)

    if progress_callback:
        progress_callback("Interrogation d'OpenStreetMap (Overpass)...")
    query = build_overpass_query(geo, tag_pairs)
    data = run_overpass(query, progress_callback=progress_callback)

    results = []
    seen = set()
    for element in data.get("elements", []):
        tags = element.get("tags", {})
        name = tags.get("name")
        if not name:
            continue  # on ignore les établissements sans nom

        website_present = has_website(tags)

        if website_present:
            if mode not in ("with_website", "both"):
                continue
            statut = "Avec site web"
        else:
            if mode not in ("no_website", "both"):
                continue
            # On exclut les établissements totalement injoignables
            if not (has_phone(tags) or has_email(tags)):
                continue
            statut = "Sans site web"

        if element["type"] == "node":
            lat, lon = element.get("lat"), element.get("lon")
        else:
            center = element.get("center", {})
            lat, lon = center.get("lat"), center.get("lon")
        if lat is None or lon is None:
            continue

        key = (name, round(lat, 5), round(lon, 5))
        if key in seen:
            continue
        seen.add(key)

        results.append(
            {
                "nom": name,
                "statut": statut,
                "adresse": extract_address(tags),
                "telephone": extract_phone(tags),
                "email": extract_email(tags),
                "site_web": extract_website(tags),
                "lat": lat,
                "lon": lon,
                "maps_url": f"https://www.google.com/maps?q={lat},{lon}",
            }
        )

    if progress_callback:
        progress_callback(f"{len(results)} établissement(s) trouvé(s).")
    return results


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------
def export_to_xlsx(results, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Établissements"

    headers = ["Nom", "Statut", "Adresse", "Téléphone", "E-mail", "Site web", "Lien Google Maps"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for r in results:
        ws.append([r["nom"], r["statut"], r["adresse"], r["telephone"], r["email"], r["site_web"], r["maps_url"]])

    widths = [28, 16, 40, 16, 26, 30, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Interface graphique (Tkinter)
# ---------------------------------------------------------------------------
class App(tk.Tk):
    SUGGEST_DEBOUNCE_MS = 450

    def __init__(self):
        super().__init__()
        self.title("Recherche d'établissements (données OpenStreetMap)")
        self.geometry("1100x620")
        self.minsize(900, 520)

        self.results = []
        self.selected_geo = None  # géocodage choisi via une suggestion
        self._suggestions = []
        self._suggest_after_id = None
        self._suggest_popup = None
        self._suggest_listbox = None

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Ville :").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.city_var = tk.StringVar()
        self.city_entry = ttk.Entry(top, textvariable=self.city_var, width=22)
        self.city_entry.grid(row=0, column=1, padx=(0, 15))
        self.city_entry.bind("<KeyRelease>", self.on_city_keyrelease)
        self.city_entry.bind("<FocusOut>", lambda e: self.after(150, self.hide_suggestions))
        self.city_entry.bind("<Escape>", lambda e: self.hide_suggestions())

        ttk.Label(top, text="Pays / État :").grid(row=0, column=2, sticky="w", padx=(0, 5))
        self.region_var = tk.StringVar()
        self.region_entry = ttk.Entry(top, textvariable=self.region_var, width=16)
        self.region_entry.grid(row=0, column=3, padx=(0, 15))
        self.region_entry.bind("<KeyRelease>", self.on_city_keyrelease)

        ttk.Label(top, text="Catégorie :").grid(row=0, column=4, sticky="w", padx=(0, 5))
        self.category_var = tk.StringVar()
        category_box = ttk.Combobox(
            top, textvariable=self.category_var, width=18,
            values=list(CATEGORIES.keys()),
        )
        category_box.grid(row=0, column=5, padx=(0, 15))
        category_box.set("Supermarché")

        ttk.Label(top, text="Type de recherche :").grid(row=1, column=0, sticky="w", padx=(0, 5), pady=(8, 0))
        self.mode_var = tk.StringVar()
        mode_box = ttk.Combobox(
            top, textvariable=self.mode_var, width=38,
            values=[MODE_SANS_SITE, MODE_AVEC_SITE, MODE_LES_DEUX],
            state="readonly",
        )
        mode_box.grid(row=1, column=1, columnspan=3, sticky="w", pady=(8, 0))
        mode_box.set(MODE_SANS_SITE)

        self.search_btn = ttk.Button(top, text="Rechercher", command=self.on_search)
        self.search_btn.grid(row=1, column=4, pady=(8, 0), padx=(15, 10))

        self.export_btn = ttk.Button(top, text="Exporter en Excel", command=self.on_export, state="disabled")
        self.export_btn.grid(row=1, column=5, pady=(8, 0))

        hint = ttk.Label(
            self, foreground="#555",
            text="Astuce : tapez la ville (et le pays) pour voir des suggestions, puis cliquez sur la bonne ville dans la liste.",
            padding=(10, 0),
        )
        hint.pack(fill="x")

        self.status_var = tk.StringVar(value="Prêt.")
        ttk.Label(self, textvariable=self.status_var, padding=(10, 0)).pack(fill="x")

        columns = ("nom", "statut", "adresse", "telephone", "email", "site_web", "maps")
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        self.tree.heading("nom", text="Nom")
        self.tree.heading("statut", text="Statut")
        self.tree.heading("adresse", text="Adresse")
        self.tree.heading("telephone", text="Téléphone")
        self.tree.heading("email", text="E-mail")
        self.tree.heading("site_web", text="Site web")
        self.tree.heading("maps", text="Lien Google Maps (double-clic)")
        self.tree.column("nom", width=160)
        self.tree.column("statut", width=110)
        self.tree.column("adresse", width=230)
        self.tree.column("telephone", width=100)
        self.tree.column("email", width=150)
        self.tree.column("site_web", width=160)
        self.tree.column("maps", width=190)
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.bind("<Double-1>", self.on_row_double_click)

        scrollbar = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    # -- Suggestions de ville en direct --------------------------------
    def on_city_keyrelease(self, event):
        # Toute frappe invalide la sélection précédente : l'utilisateur
        # doit recliquer une suggestion, ou une recherche directe sera
        # faite sur le texte tapé.
        self.selected_geo = None

        if self._suggest_after_id is not None:
            self.after_cancel(self._suggest_after_id)
        self._suggest_after_id = self.after(self.SUGGEST_DEBOUNCE_MS, self.fetch_suggestions)

    def fetch_suggestions(self):
        city = self.city_var.get().strip()
        region = self.region_var.get().strip()
        if len(city) < 2:
            self.hide_suggestions()
            return
        thread = threading.Thread(target=self._suggestions_worker, args=(city, region), daemon=True)
        thread.start()

    def _suggestions_worker(self, city, region):
        try:
            candidates = geocode_candidates(city, region, limit=6)
        except Exception:  # noqa: BLE001
            candidates = []
        self.after(0, self.show_suggestions, candidates)

    def show_suggestions(self, candidates):
        self._suggestions = candidates
        if not candidates:
            self.hide_suggestions()
            return

        if self._suggest_popup is None:
            self._suggest_popup = tk.Toplevel(self)
            self._suggest_popup.overrideredirect(True)
            self._suggest_popup.attributes("-topmost", True)
            self._suggest_listbox = tk.Listbox(self._suggest_popup, height=6, activestyle="dotbox")
            self._suggest_listbox.pack(fill="both", expand=True)
            self._suggest_listbox.bind("<<ListboxSelect>>", self.on_suggestion_select)
            self._suggest_listbox.bind("<Button-1>", self.on_suggestion_select, add="+")

        self._suggest_listbox.delete(0, "end")
        for c in candidates:
            self._suggest_listbox.insert("end", c.get("display_name", "")[:90])

        x = self.city_entry.winfo_rootx()
        y = self.city_entry.winfo_rooty() + self.city_entry.winfo_height()
        width = max(self.city_entry.winfo_width() + self.region_entry.winfo_width() + 20, 350)
        height = min(22 * len(candidates) + 6, 140)
        self._suggest_popup.geometry(f"{width}x{height}+{x}+{y}")
        self._suggest_popup.deiconify()
        self._suggest_popup.lift()

    def on_suggestion_select(self, event):
        # Utilise after(10, ...) pour laisser Tkinter enregistrer le
        # clic avant de lire la sélection courante.
        self.after(10, self._apply_suggestion_selection)

    def _apply_suggestion_selection(self):
        if self._suggest_listbox is None:
            return
        selection = self._suggest_listbox.curselection()
        if not selection or selection[0] >= len(self._suggestions):
            return
        candidate = self._suggestions[selection[0]]
        self.selected_geo = candidate
        display_name = candidate.get("display_name", "")
        # On garde le nom complet affiché pour éviter toute ambiguïté
        self.city_var.set(display_name.split(",")[0].strip())
        self.hide_suggestions()

    def hide_suggestions(self):
        if self._suggest_popup is not None:
            self._suggest_popup.withdraw()

    # -- Actions ------------------------------------------------------
    def on_search(self):
        self.hide_suggestions()
        city = self.city_var.get().strip()
        region = self.region_var.get().strip()
        category = self.category_var.get().strip()
        mode_label = self.mode_var.get().strip()

        if not city:
            messagebox.showwarning("Champ manquant", "Merci de saisir une ville.")
            return
        if not category:
            messagebox.showwarning("Champ manquant", "Merci de saisir ou choisir une catégorie.")
            return

        mode = MODE_CODES.get(mode_label, "no_website")

        self.search_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.results = []

        if self.selected_geo is not None:
            # L'utilisateur a cliqué une suggestion : on part directement,
            # sans nouvelle recherche ni fenêtre de confirmation.
            self.status_var.set("Interrogation d'OpenStreetMap...")
            self._start_overpass_search(self.selected_geo, category, mode)
        else:
            # Pas de suggestion sélectionnée : on prend directement la
            # meilleure correspondance, sans bloquer avec une fenêtre.
            self.status_var.set("Localisation de la ville...")
            thread = threading.Thread(
                target=self._geocode_worker, args=(city, region, category, mode), daemon=True
            )
            thread.start()

    def _geocode_worker(self, city, region, category, mode):
        try:
            geo = geocode_best_match(city, region)
            self.after(0, self._on_geocode_done, geo, category, mode, None)
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._on_geocode_done, None, category, mode, exc)

    def _on_geocode_done(self, geo, category, mode, error):
        if error is not None:
            messagebox.showerror("Erreur", str(error))
            self.status_var.set("Erreur lors de la localisation de la ville.")
            self.search_btn.config(state="normal")
            return
        self._start_overpass_search(geo, category, mode)

    def _start_overpass_search(self, geo, category, mode):
        self.status_var.set("Interrogation d'OpenStreetMap...")
        thread = threading.Thread(target=self._search_worker, args=(geo, category, mode), daemon=True)
        thread.start()

    def _search_worker(self, geo, category, mode):
        try:
            results = search_places(geo, category, mode=mode, progress_callback=self._set_status_threadsafe)
            self.after(0, self._on_search_done, results, None)
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._on_search_done, None, exc)

    def _set_status_threadsafe(self, message):
        self.after(0, self.status_var.set, message)

    def _on_search_done(self, results, error):
        self.search_btn.config(state="normal")
        if error is not None:
            messagebox.showerror("Erreur", str(error))
            self.status_var.set("Erreur lors de la recherche.")
            return

        self.results = results
        for r in results:
            self.tree.insert(
                "", "end",
                values=(r["nom"], r["statut"], r["adresse"], r["telephone"], r["email"], r["site_web"], r["maps_url"]),
            )

        if results:
            self.export_btn.config(state="normal")
            self.status_var.set(f"{len(results)} établissement(s) trouvé(s).")
        else:
            self.status_var.set("Aucun établissement trouvé pour cette recherche.")

    def on_row_double_click(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        values = self.tree.item(item_id, "values")
        if len(values) >= 7:
            webbrowser.open(values[6])

    def on_export(self):
        if not self.results:
            messagebox.showinfo("Rien à exporter", "Aucun résultat à exporter.")
            return
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
            initialfile="resultats_etablissements.xlsx",
        )
        if not filepath:
            return
        try:
            export_to_xlsx(self.results, filepath)
            messagebox.showinfo("Export réussi", f"Résultats exportés vers :\n{filepath}")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Erreur d'export", str(exc))


if __name__ == "__main__":
    app = App()
    app.mainloop()
